from decimal import Decimal
from datetime import date, timedelta
import json

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count, Q
from django.db.models.functions import TruncDate
from django.utils import timezone
from django.core.paginator import Paginator

from .models import GCashTransaction, GCashSettings
from .forms import GCashTransactionForm, GCashSettingsForm


def _completed_qs():
    return GCashTransaction.objects.filter(status=GCashTransaction.STATUS_COMPLETED)


@login_required
def dashboard(request):
    today  = timezone.localdate()
    cfg    = GCashSettings.get()
    period = request.GET.get('period', 'daily')

    if period == 'weekly':
        start = today - timedelta(days=today.weekday())
        end   = today
    elif period == 'monthly':
        start = today.replace(day=1)
        end   = today
    else:
        start = end = today

    qs = _completed_qs().filter(created_at__date__gte=start, created_at__date__lte=end)

    total_cash_in  = qs.filter(txn_type='cash_in').aggregate(t=Sum('amount'))['t']  or Decimal('0')
    total_cash_out = qs.filter(txn_type='cash_out').aggregate(t=Sum('amount'))['t'] or Decimal('0')
    total_fees     = qs.aggregate(t=Sum('service_fee'))['t'] or Decimal('0')
    txn_count      = qs.count()
    cash_in_count  = qs.filter(txn_type='cash_in').count()
    cash_out_count = qs.filter(txn_type='cash_out').count()

    # 30-day fee chart
    chart_labels, chart_fees = [], []
    for i in range(29, -1, -1):
        d = today - timedelta(days=i)
        day_fees = _completed_qs().filter(created_at__date=d).aggregate(t=Sum('service_fee'))['t'] or 0
        chart_labels.append(d.strftime('%b %d'))
        chart_fees.append(float(day_fees))

    recent = GCashTransaction.objects.select_related('processed_by')[:10]

    return render(request, 'gcash/dashboard.html', {
        'period':        period,
        'total_cash_in':  total_cash_in,
        'total_cash_out': total_cash_out,
        'total_fees':     total_fees,
        'txn_count':      txn_count,
        'cash_in_count':  cash_in_count,
        'cash_out_count': cash_out_count,
        'recent':         recent,
        'cfg':            cfg,
        'chart_labels':   json.dumps(chart_labels),
        'chart_fees':     json.dumps(chart_fees),
    })


@login_required
def transaction_create(request):
    cfg = GCashSettings.get()
    if request.method == 'POST':
        form = GCashTransactionForm(request.POST)
        if form.is_valid():
            txn = form.save(commit=False)
            txn.processed_by = request.user
            txn.save()
            messages.success(request, f'Transaction {txn.txn_number} recorded successfully.')
            return redirect('gcash:transaction_detail', pk=txn.pk)
    else:
        form = GCashTransactionForm()
    return render(request, 'gcash/transaction_form.html', {
        'form': form, 'cfg': cfg, 'action': 'New',
    })


@login_required
def transaction_list(request):
    qs = GCashTransaction.objects.select_related('processed_by')

    txn_type  = request.GET.get('type', '')
    status    = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    staff_id  = request.GET.get('staff', '')
    search    = request.GET.get('q', '')

    if txn_type:
        qs = qs.filter(txn_type=txn_type)
    if status:
        qs = qs.filter(status=status)
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)
    if staff_id:
        qs = qs.filter(processed_by_id=staff_id)
    if search:
        qs = qs.filter(
            Q(txn_number__icontains=search) |
            Q(customer_name__icontains=search) |
            Q(reference_number__icontains=search)
        )

    paginator = Paginator(qs, 50)
    page      = paginator.get_page(request.GET.get('page'))

    completed_qs = qs.filter(status='completed')
    totals = completed_qs.aggregate(
        total_amount=Sum('amount'),
        total_fees=Sum('service_fee'),
        count=Count('id'),
    )

    from accounts.models import User
    staff_list = User.objects.filter(is_active=True).order_by('username')

    return render(request, 'gcash/transaction_list.html', {
        'page':       page,
        'totals':     totals,
        'staff_list': staff_list,
        'filters': {
            'type': txn_type, 'status': status,
            'date_from': date_from, 'date_to': date_to,
            'staff': staff_id, 'q': search,
        },
    })


@login_required
def transaction_detail(request, pk):
    txn = get_object_or_404(GCashTransaction, pk=pk)
    return render(request, 'gcash/transaction_detail.html', {'txn': txn})


@login_required
def transaction_edit(request, pk):
    txn = get_object_or_404(GCashTransaction, pk=pk)
    if txn.status == GCashTransaction.STATUS_VOIDED:
        messages.error(request, 'Cannot edit a voided transaction.')
        return redirect('gcash:transaction_detail', pk=pk)
    if not request.user.is_admin_or_manager:
        messages.error(request, 'Only managers and admins can edit transactions.')
        return redirect('gcash:transaction_detail', pk=pk)
    cfg = GCashSettings.get()
    if request.method == 'POST':
        form = GCashTransactionForm(request.POST, instance=txn)
        if form.is_valid():
            form.save()
            messages.success(request, f'Transaction {txn.txn_number} updated.')
            return redirect('gcash:transaction_detail', pk=pk)
    else:
        form = GCashTransactionForm(instance=txn)
    return render(request, 'gcash/transaction_form.html', {
        'form': form, 'cfg': cfg, 'action': 'Edit', 'txn': txn,
    })


@login_required
def transaction_void(request, pk):
    txn = get_object_or_404(GCashTransaction, pk=pk)
    if txn.status == GCashTransaction.STATUS_VOIDED:
        messages.error(request, 'Transaction is already voided.')
        return redirect('gcash:transaction_detail', pk=pk)
    if request.method == 'POST':
        reason = request.POST.get('void_reason', '').strip()
        if not reason:
            messages.error(request, 'Please provide a void reason.')
        else:
            txn.status      = GCashTransaction.STATUS_VOIDED
            txn.void_reason = reason
            txn.voided_by   = request.user
            txn.voided_at   = timezone.now()
            txn.save()
            messages.success(request, f'Transaction {txn.txn_number} has been voided.')
            return redirect('gcash:transaction_list')
    return render(request, 'gcash/transaction_void.html', {'txn': txn})


@login_required
def reports(request):
    if not request.user.is_admin_or_manager:
        messages.error(request, 'Access restricted to managers and admins.')
        return redirect('gcash:dashboard')

    today     = timezone.localdate()
    date_from = request.GET.get('date_from', today.strftime('%Y-%m-%d'))
    date_to   = request.GET.get('date_to',   today.strftime('%Y-%m-%d'))
    txn_type  = request.GET.get('type', '')
    staff_id  = request.GET.get('staff', '')

    qs = _completed_qs().filter(
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
    )
    if txn_type:
        qs = qs.filter(txn_type=txn_type)
    if staff_id:
        qs = qs.filter(processed_by_id=staff_id)

    summary = qs.aggregate(
        total_cash_in=Sum('amount',      filter=Q(txn_type='cash_in')),
        total_cash_out=Sum('amount',     filter=Q(txn_type='cash_out')),
        total_fees=Sum('service_fee'),
        count=Count('id'),
        cash_in_count=Count('id',  filter=Q(txn_type='cash_in')),
        cash_out_count=Count('id', filter=Q(txn_type='cash_out')),
    )

    staff_breakdown = (
        qs.values('processed_by__username')
          .annotate(count=Count('id'), total_fees=Sum('service_fee'), total_amount=Sum('amount'))
          .order_by('-total_fees')
    )

    daily = (
        qs.annotate(day=TruncDate('created_at'))
          .values('day')
          .annotate(count=Count('id'), total_fees=Sum('service_fee'), total_amount=Sum('amount'))
          .order_by('day')
    )

    if request.GET.get('export') == 'excel':
        return _export_excel(qs, date_from, date_to)

    from accounts.models import User
    staff_list = User.objects.filter(is_active=True).order_by('username')

    return render(request, 'gcash/reports.html', {
        'summary':         summary,
        'staff_breakdown': staff_breakdown,
        'daily':           daily,
        'staff_list':      staff_list,
        'transactions':    qs.select_related('processed_by')[:500],
        'filters': {
            'date_from': date_from, 'date_to': date_to,
            'type': txn_type, 'staff': staff_id,
        },
    })


def _export_excel(qs, date_from, date_to):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'GCash Transactions'

    hf = Font(bold=True, color='FFFFFF')
    hb = PatternFill('solid', fgColor='0d6efd')
    ac = Alignment(horizontal='center')

    headers = ['#', 'TXN Number', 'Date & Time', 'Type', 'Amount', 'Service Fee', 'Customer', 'Reference', 'Staff']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font, cell.fill, cell.alignment = hf, hb, ac

    for row, txn in enumerate(qs.select_related('processed_by'), 2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=txn.txn_number)
        ws.cell(row=row, column=3, value=txn.created_at.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=4, value=txn.get_txn_type_display())
        ws.cell(row=row, column=5, value=float(txn.amount))
        ws.cell(row=row, column=6, value=float(txn.service_fee))
        ws.cell(row=row, column=7, value=txn.customer_name)
        ws.cell(row=row, column=8, value=txn.reference_number)
        ws.cell(row=row, column=9, value=txn.processed_by.username)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="gcash_{date_from}_{date_to}.xlsx"'
    return resp


@login_required
def gcash_settings(request):
    if not request.user.is_admin:
        messages.error(request, 'Only admins can change GCash settings.')
        return redirect('gcash:dashboard')
    cfg = GCashSettings.get()
    if request.method == 'POST':
        form = GCashSettingsForm(request.POST, instance=cfg)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.updated_by = request.user
            obj.save()
            messages.success(request, 'GCash settings updated.')
            return redirect('gcash:settings')
    else:
        form = GCashSettingsForm(instance=cfg)
    return render(request, 'gcash/settings.html', {'form': form, 'cfg': cfg})


@login_required
def fee_api(request):
    try:
        amount = Decimal(request.GET.get('amount', '0'))
        cfg    = GCashSettings.get()
        fee    = cfg.compute_fee(amount)
        return JsonResponse({'fee': str(fee), 'fee_per_1000': str(cfg.fee_per_1000)})
    except Exception:
        return JsonResponse({'fee': '0'})
