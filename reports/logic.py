"""
reports/logic.py  —  Pure data-layer functions, no HTTP.
All aggregations live here so views stay thin.
"""
from decimal import Decimal
from django.db.models import Sum, Count, Q, F
from django.db.models.functions import TruncDate
from django.utils import timezone

from transactions.models import Payment


# ──────────────────────────────────────────────────────────────
# Helper
# ──────────────────────────────────────────────────────────────

def _completed(qs):
    return qs.filter(status='completed')


def _zero(val):
    return val if val is not None else Decimal('0')


# ──────────────────────────────────────────────────────────────
# Daily summary  (single date)
# ──────────────────────────────────────────────────────────────

def daily_summary(date):
    qs = Payment.objects.filter(created_at__date=date)
    completed = _completed(qs)

    agg = completed.aggregate(
        txn_count       = Count('id'),
        total_bill      = Sum('bill_amount'),
        total_fees      = Sum('service_fee'),
        total_collected = Sum('total_due'),
        total_cash_in   = Sum('cash_received'),
        total_change    = Sum('change_given'),
    )

    return {
        'date'            : date,
        'txn_count'       : agg['txn_count']       or 0,
        'total_bill'      : _zero(agg['total_bill']),
        'total_fees'      : _zero(agg['total_fees']),
        'total_collected' : _zero(agg['total_collected']),
        'total_cash_in'   : _zero(agg['total_cash_in']),
        'total_change'    : _zero(agg['total_change']),
        'voided_count'    : qs.filter(status='voided').count(),
        'pending_count'   : qs.filter(status='pending_void').count(),
    }


# ──────────────────────────────────────────────────────────────
# Date-range summary  (for range report + chart data)
# ──────────────────────────────────────────────────────────────

def range_summary(date_from, date_to):
    qs = Payment.objects.filter(
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
    )
    completed = _completed(qs)

    agg = completed.aggregate(
        txn_count       = Count('id'),
        total_bill      = Sum('bill_amount'),
        total_fees      = Sum('service_fee'),
        total_collected = Sum('total_due'),
    )

    # Daily breakdown for the chart
    daily = (
        completed
        .annotate(day=TruncDate('created_at'))
        .values('day')
        .annotate(
            count    = Count('id'),
            collected= Sum('total_due'),
            fees     = Sum('service_fee'),
            bill     = Sum('bill_amount'),
        )
        .order_by('day')
    )

    return {
        'date_from'       : date_from,
        'date_to'         : date_to,
        'txn_count'       : agg['txn_count']       or 0,
        'total_bill'      : _zero(agg['total_bill']),
        'total_fees'      : _zero(agg['total_fees']),
        'total_collected' : _zero(agg['total_collected']),
        'voided_count'    : qs.filter(status='voided').count(),
        'daily'           : list(daily),
    }


# ──────────────────────────────────────────────────────────────
# Teller breakdown  (for a date range)
# ──────────────────────────────────────────────────────────────

def teller_breakdown(date_from, date_to):
    qs = _completed(Payment.objects.filter(
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
    ))

    rows = (
        qs
        .values('teller__id', 'teller__first_name', 'teller__last_name', 'teller__username')
        .annotate(
            count           = Count('id'),
            total_collected = Sum('total_due'),
            total_fees      = Sum('service_fee'),
            total_bill      = Sum('bill_amount'),
        )
        .order_by('-total_collected')
    )

    result = []
    for r in rows:
        fn = r['teller__first_name']
        ln = r['teller__last_name']
        name = f"{fn} {ln}".strip() or r['teller__username']
        result.append({
            'teller_id'      : r['teller__id'],
            'teller_name'    : name,
            'count'          : r['count'],
            'total_collected': _zero(r['total_collected']),
            'total_fees'     : _zero(r['total_fees']),
            'total_bill'     : _zero(r['total_bill']),
        })
    return result


# ──────────────────────────────────────────────────────────────
# Void / cancelled transactions
# ──────────────────────────────────────────────────────────────

def voided_transactions(date_from, date_to):
    return (
        Payment.objects
        .filter(
            status__in=['voided', 'pending_void'],
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,
        )
        .select_related('teller', 'void_requested_by', 'void_approved_by')
        .order_by('-created_at')
    )


# ──────────────────────────────────────────────────────────────
# Transaction detail rows  (for Excel/PDF full export)
# ──────────────────────────────────────────────────────────────

def transaction_rows(date_from, date_to, teller_id=None):
    qs = Payment.objects.filter(
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
    ).select_related('teller').order_by('created_at')

    if teller_id:
        qs = qs.filter(teller_id=teller_id)

    return qs


# ──────────────────────────────────────────────────────────────
# Excel export  (openpyxl)
# ──────────────────────────────────────────────────────────────

def build_excel(date_from, date_to, teller_id=None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Styles ──────────────────────────────────────────────
    navy   = '0F1923'
    green  = '166534'
    light  = 'F0F9FF'
    white  = 'FFFFFF'
    gray   = 'F3F4F6'
    border_color = 'D1D5DB'

    def hdr_font(color=white, bold=True, size=11):
        return Font(name='Calibri', bold=bold, color=color, size=size)

    def body_font(bold=False, color='111827', size=10):
        return Font(name='Calibri', bold=bold, color=color, size=size)

    def fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    def thin_border():
        s = Side(style='thin', color=border_color)
        return Border(left=s, right=s, top=s, bottom=s)

    peso = '#,##0.00'

    # ── Sheet 1: Summary ────────────────────────────────────
    ws = wb.active
    ws.title = 'Summary'
    ws.sheet_view.showGridLines = False

    summary = range_summary(date_from, date_to)
    teller_data = teller_breakdown(date_from, date_to)

    # Title block
    ws.merge_cells('A1:F1')
    ws['A1'] = '⚡ LEYECO Payment Monitor — Report'
    ws['A1'].font = Font(name='Calibri', bold=True, size=14, color=white)
    ws['A1'].fill = fill(navy)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:F2')
    ws['A2'] = f"Period: {date_from.strftime('%B %d, %Y')} — {date_to.strftime('%B %d, %Y')}"
    ws['A2'].font = Font(name='Calibri', size=10, color='9CA3AF')
    ws['A2'].fill = fill(navy)
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 18

    # KPI block
    kpi_row = 4
    kpis = [
        ('Total Transactions', summary['txn_count'], None),
        ('Total Collected (₱)', float(summary['total_collected']), peso),
        ('Service Fee Income (₱)', float(summary['total_fees']), peso),
        ('LEYECO Remittance (₱)', float(summary['total_bill']), peso),
        ('Voided Transactions', summary['voided_count'], None),
    ]
    for i, (label, value, fmt) in enumerate(kpis):
        col = i + 1
        lbl_cell = ws.cell(row=kpi_row,   column=col, value=label)
        val_cell = ws.cell(row=kpi_row+1, column=col, value=value)
        lbl_cell.font = Font(name='Calibri', size=9, bold=True, color='6B7280')
        lbl_cell.fill = fill(gray)
        lbl_cell.alignment = Alignment(horizontal='center')
        val_cell.font = Font(name='Calibri', size=13, bold=True, color=navy if fmt is None else green)
        val_cell.fill = fill(light if fmt else gray)
        val_cell.alignment = Alignment(horizontal='center')
        if fmt:
            val_cell.number_format = fmt
        ws.row_dimensions[kpi_row+1].height = 24

    # Teller breakdown table
    tbl_start = kpi_row + 4
    headers = ['Teller', 'Transactions', 'Total Collected (₱)', 'Service Fees (₱)', 'LEYECO Amount (₱)']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=tbl_start, column=ci, value=h)
        c.font = hdr_font()
        c.fill = fill(navy)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border()
    ws.row_dimensions[tbl_start].height = 20

    for ri, row in enumerate(teller_data, 1):
        r = tbl_start + ri
        row_fill = fill(gray) if ri % 2 == 0 else fill(white)
        vals = [row['teller_name'], row['count'],
                float(row['total_collected']), float(row['total_fees']), float(row['total_bill'])]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.font = body_font()
            c.fill = row_fill
            c.border = thin_border()
            c.alignment = Alignment(horizontal='center' if ci > 1 else 'left')
            if ci > 2:
                c.number_format = peso

    # Column widths
    col_widths = [24, 16, 22, 20, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: All Transactions ────────────────────────────
    ws2 = wb.create_sheet('Transactions')
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells('A1:J1')
    ws2['A1'] = f'All Transactions — {date_from} to {date_to}'
    ws2['A1'].font = Font(name='Calibri', bold=True, size=12, color=white)
    ws2['A1'].fill = fill(navy)
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 28

    txn_headers = [
        'Ref #', 'Date', 'Time', 'Teller', 'Biller Name',
        'Account No.', 'Bill Amount (₱)', 'Service Fee (₱)',
        'Total Due (₱)', 'Cash Received (₱)', 'Change (₱)', 'Status'
    ]
    for ci, h in enumerate(txn_headers, 1):
        c = ws2.cell(row=2, column=ci, value=h)
        c.font = hdr_font(size=9)
        c.fill = fill('1E3A5F')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border()
    ws2.row_dimensions[2].height = 32

    payments = transaction_rows(date_from, date_to, teller_id)
    for ri, p in enumerate(payments, 1):
        r = 2 + ri
        row_fill = fill(gray) if ri % 2 == 0 else fill(white)
        teller_name = p.teller.get_full_name() or p.teller.username
        vals = [
            p.txn_number,
            p.created_at.strftime('%Y-%m-%d'),
            p.created_at.strftime('%I:%M %p'),
            teller_name,
            p.biller_name,
            p.biller_account_number,
            float(p.bill_amount),
            float(p.service_fee),
            float(p.total_due),
            float(p.cash_received),
            float(p.change_given),
            p.get_status_display(),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws2.cell(row=r, column=ci, value=v)
            c.font = body_font()
            c.fill = row_fill if p.status == 'completed' else fill('FEF2F2')
            c.border = thin_border()
            c.alignment = Alignment(horizontal='center' if ci not in (1, 5, 6) else 'left')
            if ci in (7, 8, 9, 10, 11):
                c.number_format = peso

    # Totals row
    if payments.count() > 0:
        total_row = 2 + payments.count() + 1
        ws2.cell(row=total_row, column=5, value='TOTALS').font = hdr_font(color=navy)
        completed_qs = payments.filter(status='completed')
        from django.db.models import Sum as DSum
        agg = completed_qs.aggregate(
            tb=DSum('bill_amount'), tf=DSum('service_fee'),
            tc=DSum('total_due'),  tca=DSum('cash_received'), tch=DSum('change_given')
        )
        total_vals = {7: agg['tb'], 8: agg['tf'], 9: agg['tc'], 10: agg['tca'], 11: agg['tch']}
        for ci, v in total_vals.items():
            c = ws2.cell(row=total_row, column=ci, value=float(v or 0))
            c.font = Font(name='Calibri', bold=True, size=10, color=green)
            c.fill = fill('F0FDF4')
            c.number_format = peso
            c.border = thin_border()

    # Col widths for transactions sheet
    txn_widths = [22, 12, 10, 18, 24, 16, 14, 13, 13, 14, 12, 12]
    for i, w in enumerate(txn_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Voided ─────────────────────────────────────
    ws3 = wb.create_sheet('Voided')
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells('A1:G1')
    ws3['A1'] = f'Voided / Cancelled Transactions — {date_from} to {date_to}'
    ws3['A1'].font = Font(name='Calibri', bold=True, size=12, color=white)
    ws3['A1'].fill = fill('7F1D1D')
    ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 28

    void_headers = ['Ref #', 'Date', 'Biller', 'Amount (₱)', 'Requested By', 'Approved By', 'Reason']
    for ci, h in enumerate(void_headers, 1):
        c = ws3.cell(row=2, column=ci, value=h)
        c.font = hdr_font(size=9)
        c.fill = fill('991B1B')
        c.border = thin_border()
        c.alignment = Alignment(horizontal='center')

    voids = voided_transactions(date_from, date_to)
    for ri, p in enumerate(voids, 1):
        r = 2 + ri
        req = p.void_requested_by
        apv = p.void_approved_by
        vals = [
            p.txn_number,
            p.created_at.strftime('%Y-%m-%d'),
            p.biller_name,
            float(p.total_due),
            (req.get_full_name() or req.username) if req else '—',
            (apv.get_full_name() or apv.username) if apv else 'Pending',
            p.void_reason or '—',
        ]
        for ci, v in enumerate(vals, 1):
            c = ws3.cell(row=r, column=ci, value=v)
            c.font = body_font()
            c.fill = fill('FEF2F2')
            c.border = thin_border()
            if ci == 4:
                c.number_format = peso

    void_widths = [22, 12, 24, 14, 18, 18, 40]
    for i, w in enumerate(void_widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    return wb


# ──────────────────────────────────────────────────────────────
# PDF export  (pure HTML → WeasyPrint)
# ──────────────────────────────────────────────────────────────

def build_pdf_html(date_from, date_to, report_type='daily'):
    """Returns HTML string that WeasyPrint converts to PDF."""
    summary  = range_summary(date_from, date_to)
    tellers  = teller_breakdown(date_from, date_to)
    payments = transaction_rows(date_from, date_to)
    voids    = voided_transactions(date_from, date_to)
    from django.utils.timezone import now

    rows_html = ''
    for p in payments:
        teller_name = p.teller.get_full_name() or p.teller.username
        status_style = ''
        if p.status == 'voided':
            status_style = 'background:#fee2e2;'
        elif p.status == 'pending_void':
            status_style = 'background:#fef3c7;'
        rows_html += f"""
        <tr style="{status_style}">
          <td class="mono">{p.txn_number}</td>
          <td>{p.created_at.strftime('%m/%d %I:%M%p')}</td>
          <td>{teller_name}</td>
          <td>{p.biller_name}</td>
          <td>{p.biller_account_number}</td>
          <td class="num">₱{p.bill_amount:,.2f}</td>
          <td class="num fee">₱{p.service_fee:,.2f}</td>
          <td class="num bold">₱{p.total_due:,.2f}</td>
          <td class="num">₱{p.cash_received:,.2f}</td>
          <td class="num">₱{p.change_given:,.2f}</td>
          <td class="center">{p.get_status_display()}</td>
        </tr>"""

    teller_rows_html = ''
    for t in tellers:
        teller_rows_html += f"""
        <tr>
          <td>{t['teller_name']}</td>
          <td class="center">{t['count']}</td>
          <td class="num">₱{t['total_collected']:,.2f}</td>
          <td class="num fee">₱{t['total_fees']:,.2f}</td>
          <td class="num">₱{t['total_bill']:,.2f}</td>
        </tr>"""

    void_rows_html = ''
    for p in voids:
        req = p.void_requested_by
        apv = p.void_approved_by
        void_rows_html += f"""
        <tr>
          <td class="mono">{p.txn_number}</td>
          <td>{p.created_at.strftime('%m/%d')}</td>
          <td>{p.biller_name}</td>
          <td class="num">₱{p.total_due:,.2f}</td>
          <td>{(req.get_full_name() or req.username) if req else '—'}</td>
          <td>{(apv.get_full_name() or apv.username) if apv else 'Pending'}</td>
          <td>{p.void_reason or '—'}</td>
        </tr>"""

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<style>
  @page {{ size: A4 landscape; margin: 1.5cm 1cm; }}
  * {{ box-sizing: border-box; font-family: 'Arial', sans-serif; }}
  body {{ font-size: 9pt; color: #111827; }}
  .header {{ background: #0f1923; color: white; padding: 12px 16px; border-radius: 6px; margin-bottom: 16px; }}
  .header h1 {{ margin: 0; font-size: 16pt; }}
  .header p  {{ margin: 4px 0 0; font-size: 9pt; color: #9ca3af; }}
  .kpi-row {{ display: flex; gap: 12px; margin-bottom: 16px; }}
  .kpi {{ flex: 1; background: #f0f9ff; border: 1px solid #bae6fd; border-radius: 6px; padding: 10px 12px; }}
  .kpi .label {{ font-size: 8pt; color: #6b7280; font-weight: bold; text-transform: uppercase; }}
  .kpi .value {{ font-size: 14pt; font-weight: bold; color: #0f1923; margin-top: 2px; }}
  .kpi.fee .value {{ color: #166534; }}
  h2 {{ font-size: 11pt; color: #0f1923; border-bottom: 2px solid #0f1923; padding-bottom: 4px; margin: 16px 0 8px; }}
  table {{ width: 100%; border-collapse: collapse; margin-bottom: 16px; }}
  th {{ background: #0f1923; color: white; padding: 6px 8px; font-size: 8pt; text-align: center; }}
  td {{ padding: 5px 8px; border-bottom: 1px solid #e5e7eb; font-size: 8pt; }}
  tr:nth-child(even) td {{ background: #f9fafb; }}
  .num {{ text-align: right; }}
  .center {{ text-align: center; }}
  .mono {{ font-family: 'Courier New', monospace; font-size: 7.5pt; }}
  .bold {{ font-weight: bold; }}
  .fee {{ color: #166534; }}
  .total-row td {{ background: #f0fdf4 !important; font-weight: bold; border-top: 2px solid #16a34a; }}
  .footer {{ font-size: 7pt; color: #9ca3af; text-align: center; margin-top: 24px; border-top: 1px solid #e5e7eb; padding-top: 8px; }}
  .page-break {{ page-break-before: always; }}
</style>
</head>
<body>

<div class="header">
  <h1>⚡ LEYECO Payment Monitor</h1>
  <p>Report Period: {date_from.strftime('%B %d, %Y')} — {date_to.strftime('%B %d, %Y')} &nbsp;·&nbsp; Generated: {now().strftime('%B %d, %Y %I:%M %p')}</p>
</div>

<div class="kpi-row">
  <div class="kpi"><div class="label">Transactions</div><div class="value">{summary['txn_count']}</div></div>
  <div class="kpi"><div class="label">Total Collected</div><div class="value">₱{summary['total_collected']:,.2f}</div></div>
  <div class="kpi fee"><div class="label">Service Fee Income</div><div class="value">₱{summary['total_fees']:,.2f}</div></div>
  <div class="kpi"><div class="label">LEYECO Remittance</div><div class="value">₱{summary['total_bill']:,.2f}</div></div>
  <div class="kpi"><div class="label">Voided</div><div class="value">{summary['voided_count']}</div></div>
</div>

<h2>Teller Performance</h2>
<table>
  <thead><tr><th>Teller</th><th>Transactions</th><th>Total Collected</th><th>Service Fees</th><th>LEYECO Amount</th></tr></thead>
  <tbody>{teller_rows_html}</tbody>
</table>

<div class="page-break"></div>
<h2>Transaction Detail</h2>
<table>
  <thead>
    <tr>
      <th>Ref #</th><th>Date/Time</th><th>Teller</th><th>Biller</th><th>Account</th>
      <th>Bill Amount</th><th>Fee</th><th>Total</th><th>Cash In</th><th>Change</th><th>Status</th>
    </tr>
  </thead>
  <tbody>{rows_html}</tbody>
</table>

{'<div class="page-break"></div><h2>Voided / Cancelled Transactions</h2><table><thead><tr><th>Ref #</th><th>Date</th><th>Biller</th><th>Amount</th><th>Requested By</th><th>Approved By</th><th>Reason</th></tr></thead><tbody>' + void_rows_html + '</tbody></table>' if voids.exists() else ''}

<div class="footer">LEYECO Payment Monitor · Internal Use Only · Confidential</div>
</body></html>"""
