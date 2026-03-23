import io
import os
import zipfile
import csv
from datetime import date, timedelta

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone

from audit.models import AuditLog


def _admin_required(view_func):
    from functools import wraps
    @wraps(view_func)
    @login_required
    def wrapped(request, *args, **kwargs):
        if not request.user.is_admin():
            messages.error(request, 'Backup & Export is restricted to admins only.')
            return redirect('dashboard:index')
        return view_func(request, *args, **kwargs)
    return wrapped


@_admin_required
def backup_index(request):
    today = timezone.now().date()
    month_start = today.replace(day=1)

    from transactions.models import Payment
    from django.db.models import Count, Sum
    stats = {
        'total_transactions': Payment.objects.count(),
        'total_users': __import__('accounts.models', fromlist=['User']).User.objects.count() if False else _get_user_count(),
        'db_size': _get_db_size(),
    }
    return render(request, 'backup/index.html', {
        'stats': stats,
        'today': today,
        'month_start': month_start,
    })


def _get_user_count():
    from django.contrib.auth import get_user_model
    return get_user_model().objects.count()


def _get_db_size():
    """Return SQLite db file size as human-readable string."""
    from django.conf import settings
    db_path = settings.DATABASES['default'].get('NAME', '')
    try:
        size = os.path.getsize(db_path)
        if size < 1024:
            return f'{size} B'
        elif size < 1024 * 1024:
            return f'{size/1024:.1f} KB'
        else:
            return f'{size/1024/1024:.1f} MB'
    except Exception:
        return 'Unknown'


@_admin_required
def export_csv(request):
    """Export transactions to CSV for a date range."""
    today = timezone.now().date()
    date_from_str = request.GET.get('date_from', today.strftime('%Y-%m-%d'))
    date_to_str   = request.GET.get('date_to',   today.strftime('%Y-%m-%d'))

    try:
        date_from = date.fromisoformat(date_from_str)
        date_to   = date.fromisoformat(date_to_str)
    except ValueError:
        date_from = date_to = today

    from transactions.models import Payment
    payments = Payment.objects.filter(
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
    ).select_related('teller').order_by('created_at')

    response = HttpResponse(content_type='text/csv')
    filename = f'transactions_{date_from}_{date_to}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow([
        'Transaction #', 'Date', 'Time', 'Teller',
        'Biller Name', 'Account Number', 'LEYECO Reference',
        'Bill Amount', 'Service Fee', 'Total Due',
        'Cash Received', 'Change Given', 'Status', 'Notes'
    ])

    for p in payments:
        writer.writerow([
            p.txn_number,
            p.created_at.strftime('%Y-%m-%d'),
            p.created_at.strftime('%H:%M:%S'),
            p.teller.get_full_name() or p.teller.username,
            p.biller_name,
            p.biller_account_number,
            p.leyeco_reference,
            str(p.bill_amount),
            str(p.service_fee),
            str(p.total_due),
            str(p.cash_received),
            str(p.change_given),
            p.get_status_display(),
            p.notes,
        ])

    AuditLog.objects.create(
        user=request.user,
        action='updated',
        ip_address=request.META.get('REMOTE_ADDR'),
        details={
            'action_type': 'csv_export',
            'date_from': date_from_str,
            'date_to': date_to_str,
            'rows': payments.count(),
        }
    )
    return response


@_admin_required
def full_backup(request):
    """Export ALL data as a ZIP containing Excel workbooks per table."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.contrib.auth import get_user_model

    User = get_user_model()
    from transactions.models import Payment, CashDenomination
    from reports.models import EODReconciliation
    from shifts.models import Shift

    navy  = '0F1923'
    white = 'FFFFFF'
    gray  = 'F3F4F6'

    def make_sheet(wb, title, headers, rows):
        ws = wb.create_sheet(title)
        ws.sheet_view.showGridLines = False
        # Header row
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = Font(name='Calibri', bold=True, size=10, color=white)
            c.fill = PatternFill('solid', fgColor=navy)
            c.alignment = Alignment(horizontal='center')
        # Data rows
        for ri, row in enumerate(rows, 2):
            row_fill = PatternFill('solid', fgColor=gray if ri % 2 == 0 else white)
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = Font(name='Calibri', size=9)
                c.fill = row_fill
        # Auto-width (approximate)
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
        return ws

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:

        # 1. Transactions workbook
        wb_txn = openpyxl.Workbook()
        wb_txn.remove(wb_txn.active)
        payments = Payment.objects.select_related('teller').order_by('created_at')
        make_sheet(wb_txn, 'Transactions',
            ['Txn #','Date','Time','Teller','Biller','Account','LEYECO Ref',
             'Bill','Fee','Total','Cash In','Change','Status'],
            [
                (p.txn_number,
                 p.created_at.strftime('%Y-%m-%d'),
                 p.created_at.strftime('%H:%M:%S'),
                 p.teller.get_full_name() or p.teller.username,
                 p.biller_name, p.biller_account_number, p.leyeco_reference,
                 float(p.bill_amount), float(p.service_fee), float(p.total_due),
                 float(p.cash_received), float(p.change_given),
                 p.get_status_display())
                for p in payments
            ]
        )
        buf = io.BytesIO(); wb_txn.save(buf); buf.seek(0)
        zf.writestr('transactions.xlsx', buf.read())

        # 2. Users workbook
        wb_usr = openpyxl.Workbook()
        wb_usr.remove(wb_usr.active)
        make_sheet(wb_usr, 'Users',
            ['Username','Full Name','Role','Branch','Employee ID','Email',
             'Active','Date Joined','Last Login'],
            [
                (u.username, u.get_full_name(), u.get_role_display(),
                 u.branch, u.employee_id, u.email,
                 'Yes' if u.is_active else 'No',
                 u.date_joined.strftime('%Y-%m-%d') if u.date_joined else '',
                 u.last_login.strftime('%Y-%m-%d %H:%M') if u.last_login else '')
                for u in User.objects.all().order_by('username')
            ]
        )
        buf = io.BytesIO(); wb_usr.save(buf); buf.seek(0)
        zf.writestr('users.xlsx', buf.read())

        # 3. EOD workbook
        wb_eod = openpyxl.Workbook()
        wb_eod.remove(wb_eod.active)
        make_sheet(wb_eod, 'EOD Reconciliation',
            ['Date','Transactions','Total Collected','Service Fees',
             'LEYECO Amount','Declared Cash','Discrepancy','Status',
             'Closed By','Closed At'],
            [
                (str(e.date), e.total_transactions,
                 float(e.total_collected), float(e.total_service_fees),
                 float(e.total_bill_amount), float(e.declared_cash),
                 float(e.discrepancy), e.get_status_display(),
                 (e.closed_by.get_full_name() or e.closed_by.username) if e.closed_by else '',
                 e.closed_at.strftime('%Y-%m-%d %H:%M') if e.closed_at else '')
                for e in EODReconciliation.objects.all().order_by('-date')
            ]
        )
        buf = io.BytesIO(); wb_eod.save(buf); buf.seek(0)
        zf.writestr('eod_reconciliation.xlsx', buf.read())

        # 4. Shifts workbook
        wb_shft = openpyxl.Workbook()
        wb_shft.remove(wb_shft.active)
        make_sheet(wb_shft, 'Shifts',
            ['Teller','Date','Opened At','Opening Cash','Status',
             'Transactions','Collected','Fees','Expected','Declared','Discrepancy'],
            [
                (s.teller.get_full_name() or s.teller.username,
                 str(s.date), s.opened_at.strftime('%H:%M'),
                 float(s.opening_cash), s.get_status_display(),
                 s.total_transactions, float(s.total_collected),
                 float(s.total_fees), float(s.expected_cash),
                 float(s.closing_cash) if s.closing_cash else '',
                 float(s.discrepancy))
                for s in Shift.objects.select_related('teller').order_by('-opened_at')
            ]
        )
        buf = io.BytesIO(); wb_shft.save(buf); buf.seek(0)
        zf.writestr('shifts.xlsx', buf.read())

        # 5. Audit log workbook
        wb_aud = openpyxl.Workbook()
        wb_aud.remove(wb_aud.active)
        action_labels = dict(AuditLog.ACTION_CHOICES)
        make_sheet(wb_aud, 'Audit Log',
            ['Timestamp','User','Role','Action','Txn #','Biller','IP'],
            [
                (a.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                 a.user.get_full_name() or a.user.username,
                 a.user.get_role_display(),
                 action_labels.get(a.action, a.action),
                 a.target_txn.txn_number if a.target_txn else '',
                 a.target_txn.biller_name if a.target_txn else '',
                 a.ip_address or '')
                for a in AuditLog.objects.select_related('user','target_txn').order_by('-timestamp')[:10000]
            ]
        )
        buf = io.BytesIO(); wb_aud.save(buf); buf.seek(0)
        zf.writestr('audit_log.xlsx', buf.read())

        # 6. README
        ts = timezone.now().strftime('%B %d, %Y %I:%M %p')
        zf.writestr('README.txt',
            f'LEYECO Payment Monitor — Full Backup\n'
            f'Generated: {ts}\n\n'
            f'Files included:\n'
            f'  transactions.xlsx   — All payment transactions\n'
            f'  users.xlsx          — All user accounts\n'
            f'  eod_reconciliation.xlsx — All EOD records\n'
            f'  shifts.xlsx         — All teller shifts\n'
            f'  audit_log.xlsx      — Audit trail (last 10,000 entries)\n\n'
            f'Keep this file secure. It contains sensitive financial data.\n'
        )

    zip_buf.seek(0)
    filename = f'leyeco_backup_{timezone.now().strftime("%Y%m%d_%H%M")}.zip'
    response = HttpResponse(zip_buf.read(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    AuditLog.objects.create(
        user=request.user,
        action='updated',
        ip_address=request.META.get('REMOTE_ADDR'),
        details={'action_type': 'full_backup', 'filename': filename}
    )
    return response
