"""
python manage.py backup_now

Creates a full backup ZIP in the backups/ folder.
Use with Windows Task Scheduler or cron for automated daily backups.

Windows Task Scheduler example:
  Program:   python.exe
  Arguments: manage.py backup_now
  Start in:  C:/path/to/leyeco_monitor

Cron example (daily at 11 PM):
  0 23 * * * cd /path/to/leyeco_monitor && python manage.py backup_now  # daily 11pm
"""
import os
import io
import zipfile
from datetime import date

from django.core.management.base import BaseCommand
from django.utils import timezone
from django.conf import settings


class Command(BaseCommand):
    help = 'Create a full system backup ZIP in the backups/ directory'

    def add_arguments(self, parser):
        parser.add_argument(
            '--output-dir',
            default=None,
            help='Directory to save backup (default: backups/ inside project root)',
        )

    def handle(self, *args, **options):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        output_dir = options['output_dir'] or os.path.join(settings.BASE_DIR, 'backups')
        os.makedirs(output_dir, exist_ok=True)

        filename = f'leyeco_backup_{timezone.now().strftime("%Y%m%d_%H%M%S")}.zip'
        filepath = os.path.join(output_dir, filename)

        self.stdout.write(f'Creating backup: {filename}')

        # Reuse the backup logic from views
        from backup.views import full_backup as _not_used
        # Build directly here to avoid HTTP request dependency

        from transactions.models import Payment
        from django.contrib.auth import get_user_model
        from reports.models import EODReconciliation
        from shifts.models import Shift
        from audit.models import AuditLog

        User = get_user_model()
        navy  = '0F1923'
        white = 'FFFFFF'
        gray  = 'F3F4F6'

        def make_wb(title, headers, rows):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = title
            ws.sheet_view.showGridLines = False
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font = Font(name='Calibri', bold=True, size=10, color=white)
                c.fill = PatternFill('solid', fgColor=navy)
            for ri, row in enumerate(rows, 2):
                rfill = PatternFill('solid', fgColor=gray if ri%2==0 else white)
                for ci, val in enumerate(row, 1):
                    c = ws.cell(row=ri, column=ci, value=val)
                    c.font = Font(name='Calibri', size=9)
                    c.fill = rfill
            from openpyxl.utils import get_column_letter
            for col in ws.columns:
                ml = max((len(str(cell.value or '')) for cell in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(ml+4, 50)
            return wb

        with zipfile.ZipFile(filepath, 'w', zipfile.ZIP_DEFLATED) as zf:
            # Transactions
            payments = Payment.objects.select_related('teller').order_by('created_at')
            wb = make_wb('Transactions',
                ['Txn #','Date','Time','Teller','Biller','Account','Bill','Fee','Total','Status'],
                [(p.txn_number, p.created_at.strftime('%Y-%m-%d'), p.created_at.strftime('%H:%M'),
                  p.teller.get_full_name() or p.teller.username, p.biller_name,
                  p.biller_account_number, float(p.bill_amount), float(p.service_fee),
                  float(p.total_due), p.get_status_display()) for p in payments]
            )
            buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            zf.writestr('transactions.xlsx', buf.read())
            self.stdout.write(f'  ✓ transactions.xlsx ({payments.count()} rows)')

            # Users
            users = User.objects.all().order_by('username')
            wb = make_wb('Users',
                ['Username','Full Name','Role','Branch','Email','Active','Joined'],
                [(u.username, u.get_full_name(), u.get_role_display(), u.branch,
                  u.email, 'Yes' if u.is_active else 'No',
                  u.date_joined.strftime('%Y-%m-%d') if u.date_joined else '') for u in users]
            )
            buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            zf.writestr('users.xlsx', buf.read())
            self.stdout.write(f'  ✓ users.xlsx ({users.count()} rows)')

            # EOD
            eods = EODReconciliation.objects.all().order_by('-date')
            wb = make_wb('EOD',
                ['Date','Txns','Collected','Fees','Remittance','Declared','Discrepancy','Status'],
                [(str(e.date), e.total_transactions, float(e.total_collected),
                  float(e.total_service_fees), float(e.total_bill_amount),
                  float(e.declared_cash), float(e.discrepancy), e.get_status_display())
                 for e in eods]
            )
            buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            zf.writestr('eod_reconciliation.xlsx', buf.read())
            self.stdout.write(f'  ✓ eod_reconciliation.xlsx ({eods.count()} rows)')

            ts = timezone.now().strftime('%B %d, %Y %I:%M %p')
            zf.writestr('README.txt',
                f'LEYECO Backup — {ts}\n'
                f'transactions.xlsx, users.xlsx, eod_reconciliation.xlsx\n'
            )

        size = os.path.getsize(filepath)
        self.stdout.write(self.style.SUCCESS(
            f'\nBackup complete: {filepath} ({size/1024:.1f} KB)'
        ))
