"""
backup/logic.py  —  Database backup and CSV export functions.
"""
import io
import csv
import zipfile
from datetime import datetime
from django.utils import timezone


def build_full_backup():
    """
    Creates a ZIP file containing:
    - transactions.xlsx  (all payments + denominations)
    - users.xlsx         (all user accounts)
    - shifts.xlsx        (all shifts)
    - eod.xlsx           (all EOD reconciliations)
    - audit.xlsx         (audit log)
    Returns a BytesIO object.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    navy  = '0F1923'
    white = 'FFFFFF'
    gray  = 'F3F4F6'
    border_s = Side(style='thin', color='D1D5DB')
    thin = Border(left=border_s, right=border_s, top=border_s, bottom=border_s)

    def hdr(ws, row, headers, fill_hex=navy):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = Font(name='Calibri', bold=True, size=10, color=white)
            c.fill = PatternFill('solid', fgColor=fill_hex)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = thin
        ws.row_dimensions[row].height = 18

    def cell(ws, r, c, v, bold=False):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name='Calibri', size=9, bold=bold)
        cell.fill = PatternFill('solid', fgColor=gray if r % 2 == 0 else white)
        cell.border = thin
        cell.alignment = Alignment(vertical='center')
        return cell

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:

        # ── 1. Transactions ──────────────────────────────────
        from transactions.models import Payment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Payments'
        ws.sheet_view.showGridLines = False
        hdr(ws, 1, ['ID','Ref #','Date','Time','Teller','Biller Name','Account #',
                    'Bill Amount','Service Fee','Total Due','Cash Received','Change','Status',
                    'LEYECO Ref','Void Reason','Notes'])
        payments = Payment.objects.select_related('teller').order_by('created_at')
        for ri, p in enumerate(payments, 1):
            r = ri + 1
            vals = [p.pk, p.txn_number,
                    p.created_at.strftime('%Y-%m-%d'), p.created_at.strftime('%I:%M %p'),
                    p.teller.username, p.biller_name, p.biller_account_number,
                    float(p.bill_amount), float(p.service_fee), float(p.total_due),
                    float(p.cash_received), float(p.change_given), p.status,
                    p.leyeco_reference, p.void_reason, p.notes]
            for ci, v in enumerate(vals, 1):
                cell(ws, r, ci, v)
        for i, w in enumerate([6,22,12,10,16,26,16,13,12,13,14,12,12,16,20,20], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        zf.writestr('transactions.xlsx', buf.read())

        # ── 2. Users ─────────────────────────────────────────
        from django.contrib.auth import get_user_model
        User = get_user_model()
        wb2 = openpyxl.Workbook(); ws2 = wb2.active; ws2.title = 'Users'
        ws2.sheet_view.showGridLines = False
        hdr(ws2, 1, ['ID','Username','First Name','Last Name','Email','Role','Branch',
                     'Employee ID','Active','Date Joined','Last Login'])
        for ri, u in enumerate(User.objects.all().order_by('username'), 1):
            r = ri + 1
            vals = [u.pk, u.username, u.first_name, u.last_name, u.email,
                    u.role, u.branch, u.employee_id, u.is_active,
                    u.date_joined.strftime('%Y-%m-%d') if u.date_joined else '',
                    u.last_login.strftime('%Y-%m-%d %H:%M') if u.last_login else '']
            for ci, v in enumerate(vals, 1):
                cell(ws2, r, ci, v)
        for i, w in enumerate([6,16,14,14,26,12,16,12,8,14,18], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        buf2 = io.BytesIO(); wb2.save(buf2); buf2.seek(0)
        zf.writestr('users.xlsx', buf2.read())

        # ── 3. Shifts ────────────────────────────────────────
        from shifts.models import Shift
        wb3 = openpyxl.Workbook(); ws3 = wb3.active; ws3.title = 'Shifts'
        ws3.sheet_view.showGridLines = False
        hdr(ws3, 1, ['ID','Teller','Date','Status','Opening Cash','Transactions',
                     'Total Collected','Total Fees','Expected Cash','Declared Cash',
                     'Discrepancy','Opened At','Closed At','Notes'])
        for ri, s in enumerate(Shift.objects.select_related('teller').order_by('date'), 1):
            r = ri + 1
            vals = [s.pk, s.teller.username, str(s.date), s.status,
                    float(s.opening_cash), s.total_transactions,
                    float(s.total_collected), float(s.total_fees),
                    float(s.expected_cash), float(s.closing_cash or 0),
                    float(s.discrepancy),
                    s.opened_at.strftime('%Y-%m-%d %H:%M'),
                    s.closed_at.strftime('%Y-%m-%d %H:%M') if s.closed_at else '',
                    s.notes]
            for ci, v in enumerate(vals, 1):
                cell(ws3, r, ci, v)
        for i, w in enumerate([6,16,12,10,14,12,15,12,14,14,12,18,18,20], 1):
            ws3.column_dimensions[get_column_letter(i)].width = w
        buf3 = io.BytesIO(); wb3.save(buf3); buf3.seek(0)
        zf.writestr('shifts.xlsx', buf3.read())

        # ── 4. EOD Reconciliations ───────────────────────────
        from reports.models import EODReconciliation
        wb4 = openpyxl.Workbook(); ws4 = wb4.active; ws4.title = 'EOD'
        ws4.sheet_view.showGridLines = False
        hdr(ws4, 1, ['Date','Status','Transactions','Total Collected','Service Fees',
                     'Bill Amount','Declared Cash','Discrepancy','Closed By','Closed At','Notes'])
        for ri, e in enumerate(EODReconciliation.objects.select_related('closed_by').order_by('date'), 1):
            r = ri + 1
            vals = [str(e.date), e.status, e.total_transactions,
                    float(e.total_collected), float(e.total_service_fees),
                    float(e.total_bill_amount), float(e.declared_cash),
                    float(e.discrepancy),
                    e.closed_by.username if e.closed_by else '',
                    e.closed_at.strftime('%Y-%m-%d %H:%M') if e.closed_at else '',
                    e.notes]
            for ci, v in enumerate(vals, 1):
                cell(ws4, r, ci, v)
        buf4 = io.BytesIO(); wb4.save(buf4); buf4.seek(0)
        zf.writestr('eod_reconciliations.xlsx', buf4.read())

        # ── 5. Audit Log ─────────────────────────────────────
        from audit.models import AuditLog
        wb5 = openpyxl.Workbook(); ws5 = wb5.active; ws5.title = 'Audit Log'
        ws5.sheet_view.showGridLines = False
        hdr(ws5, 1, ['ID','Timestamp','User','Role','Action','Transaction Ref','IP Address','Details'])
        action_map = dict(AuditLog.ACTION_CHOICES)
        for ri, log in enumerate(AuditLog.objects.select_related('user','target_txn').order_by('timestamp'), 1):
            r = ri + 1
            vals = [log.pk,
                    log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                    log.user.username, log.user.role,
                    action_map.get(log.action, log.action),
                    log.target_txn.txn_number if log.target_txn else '',
                    log.ip_address or '',
                    str(log.details) if log.details else '']
            for ci, v in enumerate(vals, 1):
                c = cell(ws5, r, ci, v)
                if ci == 8:
                    c.alignment = Alignment(wrap_text=True, vertical='top')
        for i, w in enumerate([8,20,14,10,20,22,16,40], 1):
            ws5.column_dimensions[get_column_letter(i)].width = w
        buf5 = io.BytesIO(); wb5.save(buf5); buf5.seek(0)
        zf.writestr('audit_log.xlsx', buf5.read())

        # ── README ───────────────────────────────────────────
        now = timezone.now().strftime('%B %d, %Y %I:%M %p')
        readme = f"""LEYECO Payment Monitor — Full Database Backup
Generated: {now}

Files included:
- transactions.xlsx    All payment transactions with denominations
- users.xlsx           All user accounts (passwords not included)
- shifts.xlsx          All teller shift records
- eod_reconciliations.xlsx  All end-of-day reconciliation records
- audit_log.xlsx       Complete audit trail

IMPORTANT: Keep this file secure. It contains sensitive financial data.
Do not share with unauthorized persons.
"""
        zf.writestr('README.txt', readme)

    zip_buf.seek(0)
    return zip_buf


def build_csv_export(date_from, date_to):
    """
    Exports transactions for a date range as a simple CSV.
    Returns a BytesIO with UTF-8 CSV content.
    """
    from transactions.models import Payment
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        'Ref #', 'Date', 'Time', 'Teller', 'Biller Name', 'Account #',
        'Bill Amount', 'Service Fee', 'Total Due', 'Cash Received',
        'Change', 'Status', 'LEYECO Ref'
    ])
    payments = Payment.objects.filter(
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
    ).select_related('teller').order_by('created_at')

    for p in payments:
        writer.writerow([
            p.txn_number,
            p.created_at.strftime('%Y-%m-%d'),
            p.created_at.strftime('%I:%M %p'),
            p.teller.get_full_name() or p.teller.username,
            p.biller_name,
            p.biller_account_number,
            f'{p.bill_amount:.2f}',
            f'{p.service_fee:.2f}',
            f'{p.total_due:.2f}',
            f'{p.cash_received:.2f}',
            f'{p.change_given:.2f}',
            p.status,
            p.leyeco_reference,
        ])

    out = io.BytesIO()
    out.write(buf.getvalue().encode('utf-8-sig'))  # utf-8-sig for Excel compatibility
    out.seek(0)
    return out
