import io
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db.models import Q, Count
from django.http import HttpResponse
from django.utils import timezone
from .models import AuditLog


def _admin_required(view_func):
    from functools import wraps
    @wraps(view_func)
    @login_required
    def wrapped(request, *args, **kwargs):
        if not request.user.is_admin_or_manager():
            messages.error(request, 'Audit log access is restricted to managers and admins.')
            return redirect('dashboard:index')
        return view_func(request, *args, **kwargs)
    return wrapped


def _build_filtered_qs(request):
    """Shared filter logic used by both list and export views."""
    qs = AuditLog.objects.select_related('user', 'target_txn').order_by('-timestamp')

    action        = request.GET.get('action', '')
    user_id       = request.GET.get('user_id', '')
    date_from_str = request.GET.get('date_from', '')
    date_to_str   = request.GET.get('date_to', '')
    q             = request.GET.get('q', '').strip()

    if action:
        qs = qs.filter(action=action)
    if user_id:
        qs = qs.filter(user_id=user_id)
    if date_from_str:
        qs = qs.filter(timestamp__date__gte=date_from_str)
    if date_to_str:
        qs = qs.filter(timestamp__date__lte=date_to_str)
    if q:
        qs = qs.filter(
            Q(user__username__icontains=q) |
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q) |
            Q(target_txn__txn_number__icontains=q) |
            Q(target_txn__biller_name__icontains=q)
        )
    return qs, action, user_id, date_from_str, date_to_str, q


@_admin_required
def audit_list(request):
    qs, action, user_id, date_from_str, date_to_str, q = _build_filtered_qs(request)

    from django.contrib.auth import get_user_model
    User = get_user_model()
    all_users = User.objects.filter(is_active=True).order_by('first_name', 'username')

    # Today's counts per action type — feeds the summary pills
    today_counts = (
        AuditLog.objects.filter(timestamp__date=timezone.now().date())
        .values('action')
        .annotate(count=Count('id'))
    )
    counts = {row['action']: row['count'] for row in today_counts}

    context = {
        'logs'           : qs[:500],
        'total_count'    : qs.count(),
        'all_users'      : all_users,
        'action_choices' : AuditLog.ACTION_CHOICES,
        'filter_action'  : action,
        'filter_user'    : user_id,
        'filter_q'       : q,
        'date_from'      : date_from_str,
        'date_to'        : date_to_str,
        'today'          : timezone.now().date(),
        'action_counts'  : counts,
    }
    return render(request, 'audit/list.html', context)


@_admin_required
def export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    qs, action, user_id, date_from_str, date_to_str, q = _build_filtered_qs(request)
    logs = qs[:5000]  # cap at 5000 rows for Excel

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Audit Log'
    ws.sheet_view.showGridLines = False

    navy  = '0F1923'
    white = 'FFFFFF'
    gray  = 'F3F4F6'
    border_color = 'D1D5DB'

    def thin_border():
        s = Side(style='thin', color=border_color)
        return Border(left=s, right=s, top=s, bottom=s)

    # Title row
    ws.merge_cells('A1:F1')
    ws['A1'] = '⚡ LEYECO Monitor — Audit Log Export'
    ws['A1'].font = Font(name='Calibri', bold=True, size=13, color=white)
    ws['A1'].fill = PatternFill('solid', fgColor=navy)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Subtitle
    ws.merge_cells('A2:F2')
    filter_desc = []
    if date_from_str:
        filter_desc.append(f'From: {date_from_str}')
    if date_to_str:
        filter_desc.append(f'To: {date_to_str}')
    if action:
        filter_desc.append(f'Action: {action}')
    if q:
        filter_desc.append(f'Search: {q}')
    subtitle = '  ·  '.join(filter_desc) if filter_desc else 'All records'
    ws['A2'] = f'Filters: {subtitle}  ·  Exported: {timezone.now().strftime("%B %d, %Y %I:%M %p")}'
    ws['A2'].font = Font(name='Calibri', size=9, color='9CA3AF')
    ws['A2'].fill = PatternFill('solid', fgColor=navy)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Headers
    headers = ['Timestamp', 'User', 'Role', 'Action', 'Transaction / Detail', 'IP Address']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = Font(name='Calibri', bold=True, size=10, color=white)
        c.fill = PatternFill('solid', fgColor='1E3A5F')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border()
    ws.row_dimensions[3].height = 20

    # Action label lookup
    action_labels = dict(AuditLog.ACTION_CHOICES)

    for ri, log in enumerate(logs, 1):
        r = 3 + ri
        row_fill = PatternFill('solid', fgColor=gray if ri % 2 == 0 else white)

        teller_name = log.user.get_full_name() or log.user.username

        # Build detail string
        detail_parts = []
        if log.target_txn:
            detail_parts.append(f'{log.target_txn.txn_number} — {log.target_txn.biller_name}')
        if log.details:
            if log.details.get('reason'):
                detail_parts.append(f'Reason: {log.details["reason"][:80]}')
            if log.details.get('new_user'):
                detail_parts.append(f'Created user: {log.details["new_user"]}')
            if log.details.get('target_user'):
                detail_parts.append(f'Target user: {log.details["target_user"]}')
            if log.details.get('action_type') == 'eod_closed':
                detail_parts.append(
                    f'EOD closed · Declared: ₱{log.details.get("declared", 0)} '
                    f'· Discrepancy: ₱{log.details.get("discrepancy", 0)}'
                )
        detail = ' | '.join(detail_parts) if detail_parts else '—'

        vals = [
            log.timestamp.strftime('%Y-%m-%d %I:%M:%S %p'),
            teller_name,
            log.user.get_role_display(),
            action_labels.get(log.action, log.action),
            detail,
            log.ip_address or '—',
        ]

        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.font = Font(name='Calibri', size=9)
            c.fill = row_fill
            c.border = thin_border()
            c.alignment = Alignment(vertical='center', wrap_text=(ci == 5))

    # Column widths
    col_widths = [22, 22, 12, 20, 60, 15]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Save and return
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"audit_log_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
